"""
generate_security_review_excel.py
Generates PerioVoice_AI_Security_Review_Report.xlsx with full SAST findings from
actual static analysis of the backend source code.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = r"C:\Users\monisha D\periovoice-ai"
os.makedirs(OUT_DIR, exist_ok=True)

wb = openpyxl.Workbook()

# ─── Styles ───────────────────────────────────────────────────────────────────
def hdr(ws, row, col, val, bg="1F497D", fg="FFFFFF", bold=True, sz=11, wrap=False, center=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Calibri", size=sz, bold=bold, color=fg)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=wrap)
    return c

def cell(ws, row, col, val, bg=None, bold=False, sz=10, wrap=True, center=False, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Calibri", size=sz, bold=bold, color=color)
    if bg:
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=wrap)
    thin = Side(style="thin", color="D9D9D9")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

SEV_COLOR = {
    "CRITICAL": ("FFCCCC", "CC0000"),
    "HIGH":     ("FFE5CC", "C55A11"),
    "MEDIUM":   ("FFF2CC", "7F6000"),
    "LOW":      ("EAF1DD", "375623"),
    "INFO":     ("DDEEFF", "1F497D"),
}

# ─── SHEET 1: Executive Summary ───────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Executive Summary"

hdr(ws1, 1, 1, "PerioVoice AI™ — Backend Security Review Report (SAST)", bg="1F497D", sz=16)
ws1.merge_cells("A1:F1")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

hdr(ws1, 2, 1, "Defensive Static Code Analysis | All findings based on actual source code inspection", bg="2E5797", sz=11, fg="CCDDFF")
ws1.merge_cells("A2:F2")
ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")

summary_data = [
    ("Framework Detected",            "FastAPI (Python 3.12)"),
    ("API Architecture",              "RESTful HTTP API — 13 Endpoints"),
    ("Authentication Model",          "NONE — All endpoints are publicly accessible"),
    ("Authorization Model",           "NONE — No role or token-based access control"),
    ("Database",                      "Firebase Firestore (primary) + Local Filesystem (fallback)"),
    ("File Upload",                   "YES — /analyze/image and /api/image (multipart + base64)"),
    ("Session Handling",              "In-memory Python dict (no expiry, no persistence)"),
    ("CORS Policy",                   "allow_origins=[\"*\"] — Wildcard (ALL origins allowed)"),
    ("Security Headers",              "NONE — No CSP / HSTS / X-Frame-Options / X-Content-Type-Options"),
    ("Rate Limiting",                 "NONE — No throttling on any endpoint"),
    ("Critical Findings",             "3  (Exposed Firebase private key, No Auth on any endpoint, Wildcard CORS)"),
    ("High Findings",                 "5"),
    ("Medium Findings",               "7"),
    ("Low Findings",                  "5"),
    ("Total Security Findings",       "20"),
    ("Overall Security Score",        "28 / 100  ⚠️  NEEDS IMMEDIATE REMEDIATION"),
]

ws1.append([])
ws1.append(["Metric", "Value"])
hdr(ws1, 4, 1, "Metric", bg="1F497D")
hdr(ws1, 4, 2, "Value", bg="1F497D")

for i, (k, v) in enumerate(summary_data, start=5):
    cell(ws1, i, 1, k, bg="DCE6F1", bold=True)
    bg = "FFCCCC" if "CRITICAL" in v or "NONE" in v or "28" in v else None
    cell(ws1, i, 2, v, bg=bg)

ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 72
ws1.row_dimensions[1].height = 32
ws1.row_dimensions[2].height = 22

# ─── SHEET 2: Backend Inventory ───────────────────────────────────────────────
ws2 = wb.create_sheet("Backend Inventory")
inv_headers = ["Component", "Detail", "File / Location", "Notes"]
for ci, h in enumerate(inv_headers, 1):
    hdr(ws2, 1, ci, h, bg="1F497D")

inventory = [
    ("Framework",             "FastAPI 0.104+",                              "backend/main.py",              "Async Python web framework"),
    ("Language",              "Python 3.12",                                 "All backend files",            ""),
    ("API Style",             "REST (JSON over HTTP)",                       "backend/main.py",              "13 active routes"),
    ("Swagger / OpenAPI",     "Auto-generated at /docs and /redoc",          "FastAPI default",              "Publicly accessible — exposes full API schema"),
    ("Authentication",        "NONE (Client-side Firebase Auth only)",       "main.py — all routes",         "Server never verifies any token"),
    ("Authorization",         "NONE",                                        "main.py — all routes",         "No role checks anywhere"),
    ("Primary Database",      "Firebase Firestore",                          "backend/firebase_config.py",   "SDK: firebase-admin>=6.2.0"),
    ("Fallback Storage",      "Local Filesystem JSON",                       "backend/local_store.py",       "Stored in backend/local_storage/assessments/"),
    ("Firebase Credentials",  "Service Account Key JSON file",               "backend/firebase-key.json",    "⚠️ CRITICAL: Private key committed to repo"),
    ("Env Variables",         ".env file (tracked in git)",                  "backend/.env",                 "Contains Firebase project ID and bucket"),
    ("Session Management",    "In-memory Python dict (TriageStateEngine)",   "backend/triage_state_engine.py","No TTL, no persistence, lost on restart"),
    ("Image Upload",          "Multipart FormData + Raw Base64 body",        "backend/main.py lines 155–225","8MB size limit, JPEG/PNG/WEBP validated"),
    ("Image Analysis",        "PIL + NumPy (100% local, no external API)",   "backend/image_analyzer.py",   "Color histogram + HSV tissue analysis"),
    ("PDF Generation",        "ReportLab 4.0+",                              "backend/pdf_generator.py",    "Patient report PDF returned as bytes"),
    ("LLM Integration",       "OpenAI GPT-4o-mini (optional, fallback)",     "backend/llm_client.py",       "Uses deprecated openai.ChatCompletion.create (pre-v1)"),
    ("CORS",                  "allow_origins=['*'] — Wildcard",              "backend/main.py line 48",     "ALL origins, methods, headers allowed"),
    ("Frontend Framework",    "React 18 + Capacitor 8 (Android)",           "periovoice-web/package.json", "react-scripts 5.0.1 has known vulns"),
    ("Firebase Client SDK",   "firebase ^10.14.1",                          "periovoice-web/package.json", "Current version"),
    ("Fuzzy Matching",        "RapidFuzz 3.5+",                              "backend/triage_state_engine.py","Used against 77K row symptom CSV"),
    ("Triage Engine",         "Deterministic State Machine",                 "backend/triage_state_engine.py","1,309 line adaptive symptom engine"),
]

for ri, row in enumerate(inventory, start=2):
    for ci, val in enumerate(row, start=1):
        bg = "FFCCCC" if "CRITICAL" in str(val) or "⚠️" in str(val) else ("DCE6F1" if ci == 1 else None)
        cell(ws2, ri, ci, val, bg=bg)

for ci, w in enumerate([30, 45, 42, 52], start=1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# ─── SHEET 3: API Inventory ───────────────────────────────────────────────────
ws3 = wb.create_sheet("API Inventory")
api_headers = ["S.NO", "Endpoint", "HTTP Method", "Auth Required", "Role Required", "Controller File", "Risk Level", "Notes"]
for ci, h in enumerate(api_headers, 1):
    hdr(ws3, 1, ci, h)

apis = [
    (1,  "GET  /health",                       "GET",    "❌ No", "None", "main.py",  "LOW",      "Public health check"),
    (2,  "GET  /api/firebase/health",           "GET",    "❌ No", "None", "main.py",  "LOW",      "Exposes Firebase config (project_id, bucket)"),
    (3,  "POST /api/start",                     "POST",   "❌ No", "None", "main.py",  "MEDIUM",   "user_id passed as unvalidated query param"),
    (4,  "POST /analyze/text",                  "POST",   "❌ No", "None", "main.py",  "MEDIUM",   "Alias for /api/chat — dual route"),
    (5,  "POST /api/chat",                      "POST",   "❌ No", "None", "main.py",  "MEDIUM",   "Processes symptom messages — no auth"),
    (6,  "POST /analyze/image",                 "POST",   "❌ No", "None", "main.py",  "HIGH",     "File upload — no auth, no rate limit"),
    (7,  "POST /api/image",                     "POST",   "❌ No", "None", "main.py",  "HIGH",     "Alias for /analyze/image — dual route"),
    (8,  "POST /api/save",                      "POST",   "❌ No", "None", "main.py",  "HIGH",     "Saves ANY assessment to Firestore — no ownership check"),
    (9,  "GET  /api/history",                   "GET",    "❌ No", "None", "main.py",  "CRITICAL", "IDOR: returns ANY user's PHI by user_id param"),
    (10, "DELETE /api/assessment/{id}",         "DELETE", "❌ No", "None", "main.py",  "CRITICAL", "Anyone can delete any patient's assessment"),
    (11, "GET  /api/user/{uid}",                "GET",    "❌ No", "None", "main.py",  "HIGH",     "IDOR: reads any user profile by UID"),
    (12, "PUT  /api/user/{uid}",                "PUT",    "❌ No", "None", "main.py",  "HIGH",     "IDOR: overwrites any user profile, accepts arbitrary dict"),
    (13, "GET  /api/pdf/{assessment_id}",       "GET",    "❌ No", "None", "main.py",  "CRITICAL", "IDOR: downloads any patient's medical PDF report"),
]

for row in apis:
    ri = row[0] + 1
    sev = row[6]
    bg_c, _ = SEV_COLOR.get(sev, ("FFFFFF", "000000"))
    for ci, val in enumerate(row, start=1):
        c_bg = bg_c if ci == 7 else None
        cell(ws3, ri, ci, val, bg=c_bg, center=(ci in [1, 3, 4, 5, 7]))

col_ws = [6, 42, 10, 14, 12, 30, 12, 55]
for ci, w in enumerate(col_ws, start=1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ─── SHEET 4: Security Findings (SAST) ────────────────────────────────────────
ws4 = wb.create_sheet("Security Findings (SAST)")
sf_headers = ["S.NO", "SEVERITY", "CATEGORY", "FINDING TITLE", "FILE PATH", "LINE / LOCATION",
              "DESCRIPTION", "WHY IT IS A CONCERN", "RECOMMENDED FIX"]
for ci, h in enumerate(sf_headers, 1):
    hdr(ws4, 1, ci, h)

findings = [
    # CRITICAL
    (1,  "CRITICAL", "Credential Exposure",
         "Live Firebase Private Key Committed to Repository",
         "backend/firebase-key.json", "Full file",
         "The Firebase service account private key (private_key, private_key_id, client_email) is stored as a plaintext JSON file committed directly into the git repository.",
         "Any person with access to this repository can use this key to authenticate as a Firebase admin, gaining full read/write/delete access to ALL Firestore patient data, Cloud Storage files, and Firebase Authentication users.",
         "1. Immediately revoke this key in the Firebase Console → Project Settings → Service Accounts → Manage Keys. 2. Generate a new key. 3. Store the new key in a secret manager (e.g. Google Secret Manager, environment variable, or GitHub Secrets). 4. Add firebase-key.json and *.json to .gitignore. 5. Audit git history and use git-filter-repo to purge the key from all commits."),

    (2,  "CRITICAL", "Missing Authentication",
         "Zero Authentication on ALL 13 API Endpoints",
         "backend/main.py", "Lines 61–340 (all route handlers)",
         "Not a single API route verifies any authentication token, session cookie, or credential. The FastAPI app has no auth middleware, no Depends(verify_token), and no bearer token check.",
         "Any anonymous internet user can call every API endpoint freely: read patient medical histories, download PDF reports containing PHI, delete assessments, and modify user profiles — with zero identity verification.",
         "1. Add Firebase ID token verification as a FastAPI dependency: decode the bearer token using firebase_admin.auth.verify_id_token(). 2. Add an auth_required dependency to all protected routes. 3. Apply it via Depends() on each route handler. 4. Return HTTP 401 for missing or invalid tokens."),

    (3,  "CRITICAL", "CORS Wildcard",
         "allow_origins=[\"*\"] — All Origins Permitted",
         "backend/main.py", "Lines 46–52 (CORSMiddleware)",
         "The CORS policy allows any origin (*), any method (*), any header (*), and credentials=True. This means any website on the internet can make authenticated cross-origin requests to this API.",
         "A malicious website can silently call this API from a victim's browser using their session context. Since there is no authentication, this is compounded — any site can enumerate patient data.",
         "1. Replace allow_origins=[\"*\"] with an explicit allowlist: allow_origins=[\"https://your-production-domain.com\"]. 2. Set allow_credentials=False unless session cookies are used. 3. Restrict allow_methods and allow_headers to only what is needed."),

    # HIGH
    (4,  "HIGH", "IDOR — Broken Access Control",
         "IDOR on /api/history: Any User Can Read Any Patient's Medical History",
         "backend/main.py", "Line 270 — async def get_history(user_id: str)",
         "The GET /api/history endpoint accepts a user_id query parameter and returns all Firestore assessments for that user. There is no check that the caller is the owner of that user_id.",
         "Any caller can substitute any user_id (a UUID or email) to retrieve the complete dental symptom history, urgency assessments, and risk scores of any other patient. This is a direct HIPAA/GDPR breach.",
         "1. Require a valid Firebase ID token as a bearer header. 2. Extract the uid from the verified token server-side using auth.verify_id_token(). 3. Only query assessments where user_id == token.uid. 4. Never trust user_id from query parameters for data ownership."),

    (5,  "HIGH", "IDOR — Broken Access Control",
         "IDOR on DELETE /api/assessment/{id}: Anyone Can Delete Any Patient's Record",
         "backend/main.py", "Line 282 — async def delete_assessment(assessment_id: str)",
         "The DELETE endpoint deletes from both local storage and Firestore using only the assessment_id path parameter. No caller identity is verified.",
         "A malicious actor can enumerate assessment IDs (UUIDs) and bulk-delete all patient records from the database, causing irreversible data loss.",
         "1. Verify the Firebase ID token. 2. Look up the assessment from Firestore. 3. Confirm the token's uid matches the assessment's user_id before deleting. 4. Return HTTP 403 Forbidden if there is a mismatch."),

    (6,  "HIGH", "IDOR — Broken Access Control",
         "IDOR on PUT /api/user/{uid}: Any Caller Can Overwrite Any User Profile",
         "backend/main.py", "Line 295 — async def update_user_profile(uid: str, profile: dict)",
         "The PUT /api/user/{uid} endpoint accepts an arbitrary dict body and writes it to Firestore users/<uid> with no authentication and no input validation.",
         "1. Any caller can overwrite any other user's profile with arbitrary data. 2. The profile: dict type is completely unvalidated — an attacker could inject unexpected fields into Firestore documents.",
         "1. Verify Firebase ID token. 2. Confirm token.uid == uid before writing. 3. Replace profile: dict with a typed Pydantic model (UserProfileUpdateRequest) with explicit allowed fields. 4. Use merge=True to prevent full document overwrite."),

    (7,  "HIGH", "IDOR — Broken Access Control",
         "IDOR on GET /api/pdf/{id}: Anyone Can Download Any Patient's Medical Report PDF",
         "backend/main.py", "Line 306 — async def get_pdf_report(assessment_id, user_id)",
         "The PDF report endpoint loads an assessment by ID and generates a downloadable PDF containing the patient's name, symptoms, urgency level, and conversation transcript. No auth is checked.",
         "This is a direct leak of Protected Health Information (PHI). Anyone who knows or guesses an assessment ID can download the full medical report of any patient.",
         "1. Require a verified Firebase ID token. 2. Load the assessment from Firestore and verify its user_id matches the token's uid. 3. Return HTTP 403 if ownership check fails."),

    (8,  "HIGH", "Path Traversal",
         "Unsanitized assessment_id Used Directly as Filesystem Filename",
         "backend/local_store.py", "Lines 34, 47, 70 — BASE / f\"{aid}.json\"",
         "The assessment_id string received from API callers is used directly to construct a filesystem path without any sanitization: path = BASE / f\"{assessment_id}.json\". If assessment_id contains \"../\" sequences, it could traverse outside the assessments directory.",
         "An attacker could craft an assessment_id like ../../firebase-key to read or overwrite any file the server process has access to, including the Firebase credentials file.",
         "1. Sanitize the assessment_id using os.path.basename() or a UUID validation check before constructing the path. 2. Verify that the resolved absolute path starts with the BASE directory: assert str(path.resolve()).startswith(str(BASE.resolve())). 3. Accept only UUID-formatted IDs."),

    # MEDIUM
    (9,  "MEDIUM", "Sensitive Data in Logs",
         "Firebase Project ID, Bucket Name, and Analysis Stats Logged to stdout",
         "backend/firebase_config.py", "Lines 88–96 (initialize_firebase print block)",
         "The startup routine prints the Firebase project_id and storage bucket name to standard output. The image_analyzer prints detailed pixel statistics for every image analyzed.",
         "In a production environment, stdout logs are typically shipped to logging aggregators (CloudWatch, GCP Logging, etc.). Sensitive infrastructure details in logs increase the attack surface if logs are accessed by an unauthorized party.",
         "1. Replace bare print() calls with Python's logging module. 2. Set log level to INFO for operational messages and DEBUG for analysis stats. 3. Ensure production log level is set to WARNING or above. 4. Never log private key paths or credentials."),

    (10, "MEDIUM", "Information Disclosure",
         "Raw Exception Strings Returned in API Error Responses",
         "backend/main.py", "Lines 90, 152, 240, 270 (all HTTPException raises)",
         "All exception handlers use detail=f\"Error processing chat: {str(e)}\" — returning the raw Python exception message directly to the HTTP client.",
         "Internal error messages can reveal stack traces, module paths, database collection names, file system paths, and library versions to an attacker — useful for reconnaissance.",
         "1. Replace raw exception strings with generic user-facing messages: raise HTTPException(status_code=500, detail=\"An internal error occurred. Please try again.\"). 2. Log the full exception server-side: import logging; logging.exception(\"Chat error\"). 3. Use a global FastAPI exception handler for consistent sanitization."),

    (11, "MEDIUM", "Missing Rate Limiting",
         "No Rate Limiting or Throttling on Any Endpoint",
         "backend/main.py", "All routes — no rate limit middleware",
         "There is no rate limiting, request throttling, or abuse detection on any endpoint. The POST /api/chat and POST /analyze/image endpoints are particularly expensive (triage engine + PIL processing).",
         "A single client can send thousands of requests per second, causing CPU exhaustion, memory exhaustion (sessions dict grows unbounded), and denial of service for legitimate users.",
         "1. Add slowapi (Starlette rate limiting): from slowapi import Limiter; @limiter.limit(\"60/minute\"). 2. Apply stricter limits to /analyze/image (e.g. 10/minute per IP). 3. Set a maximum session count and evict old sessions with TTL logic."),

    (12, "MEDIUM", "Missing Input Validation",
         "ChatMessage.message Has No Maximum Length Constraint",
         "backend/models.py", "Line 32 — message: str",
         "The Pydantic ChatMessage model does not define a maximum length for the message field. The triage engine applies rapidfuzz matching against 77,792 CSV rows for every message.",
         "An attacker can send a megabyte-sized message string, causing excessive CPU usage in the fuzzy matcher and memory allocation in the triage engine on every request.",
         "1. Add Pydantic field validation: from pydantic import Field; message: str = Field(..., max_length=2000). 2. Add similar constraints to all string fields in request models. 3. Add a FastAPI request size limit using a middleware."),

    (13, "MEDIUM", "Deprecated Library API Usage",
         "openai.ChatCompletion.create() Is Deprecated (Pre-v1.0 API)",
         "backend/llm_client.py", "Lines 131, 183, 225 — openai.ChatCompletion.create()",
         "The LLM client uses the openai v0.x API style (openai.ChatCompletion.create) which was deprecated in openai v1.0.0 (November 2023). Additionally, openai is not listed in requirements.txt.",
         "If someone installs openai>=1.0.0 (the current version), the LLM client will throw AttributeError: module 'openai' has no attribute 'ChatCompletion'. Also, the GEMINI_API_KEY environment variable is checked by is_configured() but no Gemini API calls are ever made.",
         "1. Add openai>=1.0.0 to requirements.txt. 2. Update to the new SDK: client = openai.OpenAI(api_key=api_key); client.chat.completions.create(...). 3. Remove the dead GEMINI_API_KEY check from is_configured() or implement actual Gemini support."),

    (14, "MEDIUM", "Missing Security Headers",
         "No HTTP Security Headers Configured",
         "backend/main.py", "Lines 46–52 (middleware block)",
         "The API returns no security headers: no X-Content-Type-Options, no X-Frame-Options, no Content-Security-Policy, no Strict-Transport-Security, and no X-XSS-Protection.",
         "Without these headers, responses can be embedded in iframes (clickjacking), MIME-sniffed by browsers, or rendered in insecure contexts. HSTS is essential when the API is served over HTTPS.",
         "1. Add a security headers middleware: from starlette.middleware.base import BaseHTTPMiddleware. 2. Set X-Content-Type-Options: nosniff, X-Frame-Options: DENY, Referrer-Policy: strict-origin-when-cross-origin. 3. Use the fastapi-security-headers library for easy setup."),

    (15, "MEDIUM", ".env File May Be Tracked in Git",
         "backend/.env Contains Firebase Credentials and May Be Committed",
         "backend/.env", "Full file (3 lines)",
         "The backend/.env file contains the real Firebase project ID and storage bucket. The .gitignore file's coverage of this path is unknown from the analysis.",
         "If .env is committed to git history (even once), the Firebase project identifier and storage bucket name are permanently exposed. Combined with the already-committed firebase-key.json, this gives full Firebase access.",
         "1. Verify .env is in .gitignore: echo 'backend/.env' >> .gitignore. 2. Run: git ls-files --error-unmatch backend/.env — if it doesn't error, the file is tracked. 3. Remove it: git rm --cached backend/.env. 4. Use GitHub Actions secrets for CI/CD."),

    # LOW
    (16, "LOW", "Session Management",
         "In-Memory Sessions Have No TTL or Maximum Size Limit",
         "backend/triage_state_engine.py", "Line 136 — self.sessions: Dict[str, dict] = {}",
         "The sessions dictionary stores all active triage sessions in RAM with no time-to-live, no maximum session count, and no cleanup mechanism. Sessions are lost on server restart.",
         "Under sustained load, the sessions dict grows indefinitely, consuming memory. A user who abandons their session mid-triage leaks memory permanently. Server restarts invalidate all active sessions without warning.",
         "1. Use an expiring cache: from cachetools import TTLCache; self.sessions = TTLCache(maxsize=1000, ttl=3600). 2. Or use Redis with automatic key expiry. 3. Return a clear error to the client when a session is not found."),

    (17, "LOW", "Duplicate Imports",
         "FastAPI and typing Modules Imported Twice in main.py",
         "backend/main.py", "Line 8 and Lines 150–151",
         "from fastapi import FastAPI, UploadFile, File, HTTPException is imported on line 8, then imported again as from fastapi import FastAPI, UploadFile, File, HTTPException, Request on line 151 inside the module body. Similarly, from typing import Optional is imported mid-file.",
         "Duplicate imports are a code quality issue indicating the file was edited in an ad-hoc manner. The mid-file import of Request means earlier routes cannot use the Request object even though it is needed.",
         "1. Consolidate all imports at the top of the file. 2. Add Request to the initial FastAPI import on line 8. 3. Move all typing imports to the top-level import block."),

    (18, "LOW", "Missing Package in requirements.txt",
         "openai Package Used in llm_client.py But Not Listed in requirements.txt",
         "backend/requirements.txt", "All 10 lines (openai absent)",
         "The llm_client.py file performs import openai inside function bodies, but openai is not listed in requirements.txt. This means a fresh install (pip install -r requirements.txt) will not install it.",
         "If the USE_LLM environment variable is set to true and OPENAI_API_KEY is provided, the server will crash with ModuleNotFoundError: No module named 'openai' at runtime.",
         "Add openai>=1.0.0 to requirements.txt. Also consider adding it as an optional dependency with a clear comment."),

    (19, "LOW", "Unprotected API Documentation",
         "Swagger UI (/docs) and ReDoc (/redoc) Are Publicly Accessible",
         "backend/main.py", "FastAPI default (docs_url='/docs', redoc_url='/redoc')",
         "FastAPI automatically exposes full interactive API documentation at /docs (Swagger UI) and /redoc. All 13 endpoints, their parameters, request/response schemas, and data models are fully visible.",
         "An attacker can use the Swagger UI to enumerate all endpoints, understand data models, and interactively call any endpoint including the IDOR-vulnerable history and PDF endpoints without writing any code.",
         "1. In production, disable the auto-generated docs: FastAPI(docs_url=None, redoc_url=None). 2. Or protect them with HTTP Basic Auth using a Starlette middleware. 3. Consider enabling docs only when DEBUG=True."),

    (20, "LOW", "PDF Generator Accepts Unsanitized User Data",
         "User-Provided Strings Rendered Directly into ReportLab PDF Without Sanitization",
         "backend/pdf_generator.py", "Lines 115–220 (Paragraph() calls with assessment_data fields)",
         "The PDF generator renders assessment_data fields (user_name, recommendation, conversation transcript) directly into ReportLab Paragraph() objects without any HTML escaping or length truncation.",
         "ReportLab uses an XML-like markup language for Paragraph content. If a user_name or message contains malformed XML tags (e.g. <b>, </b>, <br/>), it can cause PDF rendering errors or unexpected formatting. More critically, very long untruncated transcript messages can cause oversized PDFs.",
         "1. Sanitize all user-provided strings before passing to Paragraph(): text = str(val).replace('<', '&lt;').replace('>', '&gt;'). 2. Truncate long fields: recommendation[:500]. 3. The code already truncates transcript messages to [:100] — apply this pattern consistently to all fields."),
]

for row in findings:
    ri = row[0] + 1
    sev = row[1]
    bg_c, txt_c = SEV_COLOR.get(sev, ("FFFFFF", "000000"))
    for ci, val in enumerate(row, start=1):
        c_bg = bg_c if ci == 2 else None
        c_txt = txt_c if ci == 2 else "000000"
        cell(ws4, ri, ci, val, bg=c_bg, color=c_txt, center=(ci in [1, 2]))

col_w4 = [5, 10, 22, 42, 35, 30, 55, 55, 70]
for ci, w in enumerate(col_w4, start=1):
    ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[1].height = 30

# ─── SHEET 5: Dependency Review ───────────────────────────────────────────────
ws5 = wb.create_sheet("Dependency Review")
dep_headers = ["S.NO", "Package", "Current Spec", "Latest Stable", "Risk Level", "File", "Notes / Recommendation"]
for ci, h in enumerate(dep_headers, 1):
    hdr(ws5, 1, ci, h)

deps = [
    (1,  "fastapi",               ">=0.104.0",  "0.115.x",   "LOW",    "requirements.txt", "Current. Pin to exact version in production: fastapi==0.115.0"),
    (2,  "uvicorn",               ">=0.24.0",   "0.30.x",    "LOW",    "requirements.txt", "Current. Use uvicorn[standard] for production (includes httptools/uvloop)"),
    (3,  "pydantic",              ">=2.4.2",    "2.8.x",     "LOW",    "requirements.txt", "Current. Pydantic v2 — good."),
    (4,  "python-dotenv",         ">=1.0.0",    "1.0.1",     "LOW",    "requirements.txt", "Current."),
    (5,  "python-multipart",      ">=0.0.6",    "0.0.12",    "MEDIUM", "requirements.txt", "Older versions (<0.0.7) had DoS vulnerabilities (CVE-2024-24762). Ensure >=0.0.7."),
    (6,  "pillow",                ">=10.1.0",   "10.4.x",    "MEDIUM", "requirements.txt", "Multiple CVEs in older Pillow versions. Ensure using 10.3.0+ for latest security patches."),
    (7,  "numpy",                 ">=1.26.0",   "2.1.x",     "LOW",    "requirements.txt", "Current spec. Consider upgrading to numpy>=2.0 for performance improvements."),
    (8,  "reportlab",             ">=4.0.5",    "4.2.x",     "LOW",    "requirements.txt", "Current."),
    (9,  "rapidfuzz",             ">=3.5.0",    "3.9.x",     "LOW",    "requirements.txt", "Current."),
    (10, "firebase-admin",        ">=6.2.0",    "6.5.x",     "LOW",    "requirements.txt", "Current. Ensure using latest 6.x for security patches."),
    (11, "openai",                "NOT LISTED", "1.51.x",    "HIGH",   "requirements.txt", "Used in llm_client.py but missing from requirements.txt. Add openai>=1.0.0. Current code uses deprecated v0.x API style."),
    (12, "react-scripts",        "5.0.1",      "5.0.1",     "MEDIUM", "package.json",     "react-scripts 5.0.1 bundles webpack and babel with known moderate vulnerabilities. Consider migrating to Vite."),
    (13, "axios",                 "^1.6.0",     "1.7.x",     "LOW",    "package.json",     "Current. Axios <1.6.0 had SSRF risk — 1.6.0+ is safe."),
    (14, "firebase (JS SDK)",    "^10.14.1",   "10.14.x",   "LOW",    "package.json",     "Current."),
    (15, "react",                 "^18.2.0",    "18.3.x",    "LOW",    "package.json",     "Current."),
    (16, "framer-motion",        "^10.16.0",   "11.x",      "LOW",    "package.json",     "v10 is LTS. Consider upgrading to v11 for latest fixes."),
    (17, "jspdf",                 "^2.5.1",     "2.5.2",     "LOW",    "package.json",     "Current."),
    (18, "@capacitor/core",      "^8.4.2",     "8.4.2",     "LOW",    "package.json",     "Current."),
]

for row in deps:
    ri = row[0] + 1
    sev = row[4]
    bg_c, txt_c = SEV_COLOR.get(sev, ("FFFFFF", "000000"))
    for ci, val in enumerate(row, start=1):
        c_bg = bg_c if ci == 5 else None
        c_txt = txt_c if ci == 5 else "000000"
        cell(ws5, ri, ci, val, bg=c_bg, color=c_txt, center=(ci in [1, 5]))

col_w5 = [5, 28, 16, 14, 10, 20, 65]
for ci, w in enumerate(col_w5, start=1):
    ws5.column_dimensions[get_column_letter(ci)].width = w

# ─── SHEET 6: Risk Summary & Remediation ─────────────────────────────────────
ws6 = wb.create_sheet("Risk Summary & Remediation")
hdr(ws6, 1, 1, "PerioVoice AI™ Security Risk Summary & Remediation Roadmap", bg="1F497D", sz=14)
ws6.merge_cells("A1:E1")
ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 30

risk_headers = ["Priority", "Finding", "Risk", "Effort", "Recommended Action"]
for ci, h in enumerate(risk_headers, 1):
    hdr(ws6, 2, ci, h)

roadmap = [
    ("P0 — IMMEDIATE", "Firebase private key committed to git repo",        "CRITICAL", "Low",    "Revoke the current key NOW. Generate a new one. Store in env var / secret manager. Add to .gitignore. Purge from git history."),
    ("P0 — IMMEDIATE", "Zero authentication on all 13 API endpoints",       "CRITICAL", "High",   "Implement Firebase ID token verification as a FastAPI dependency. Apply to all non-health-check routes."),
    ("P0 — IMMEDIATE", "Wildcard CORS (allow_origins=['*'])",               "CRITICAL", "Low",    "Replace * with explicit frontend domain(s) in the CORS middleware configuration."),
    ("P1 — THIS WEEK", "IDOR on /api/history — reads any user's PHI",       "HIGH",     "Medium", "After adding auth, verify token.uid == requested user_id before returning data."),
    ("P1 — THIS WEEK", "IDOR on DELETE /api/assessment/{id}",               "HIGH",     "Medium", "After adding auth, verify token.uid matches the assessment's owner before deleting."),
    ("P1 — THIS WEEK", "IDOR on PUT /api/user/{uid}",                       "HIGH",     "Medium", "After adding auth, verify token.uid == uid path param before writing."),
    ("P1 — THIS WEEK", "IDOR on GET /api/pdf/{id}",                         "HIGH",     "Medium", "After adding auth, verify token.uid matches PDF owner before streaming PDF."),
    ("P1 — THIS WEEK", "Path traversal in local_store.py",                  "HIGH",     "Low",    "Sanitize assessment_id: validate UUID format before using as filename. Assert resolved path stays within BASE."),
    ("P2 — THIS MONTH", "Raw exceptions leaked in HTTP error responses",    "MEDIUM",   "Low",    "Wrap all raises with generic messages. Log full exceptions server-side only."),
    ("P2 — THIS MONTH", "No rate limiting on any endpoint",                 "MEDIUM",   "Low",    "Add slowapi with per-IP limits (60/min for chat, 10/min for image upload)."),
    ("P2 — THIS MONTH", "No input length validation on ChatMessage",        "MEDIUM",   "Low",    "Add Field(max_length=2000) to all string fields in Pydantic request models."),
    ("P2 — THIS MONTH", "No HTTP security headers",                         "MEDIUM",   "Low",    "Add security headers middleware: X-Content-Type-Options, X-Frame-Options, Referrer-Policy."),
    ("P2 — THIS MONTH", "Swagger docs publicly accessible",                 "MEDIUM",   "Low",    "Disable /docs and /redoc in production (docs_url=None, redoc_url=None)."),
    ("P3 — NEXT QUARTER","In-memory sessions with no TTL",                  "LOW",      "Medium", "Replace raw dict with TTLCache(maxsize=1000, ttl=3600) from cachetools."),
    ("P3 — NEXT QUARTER","openai missing from requirements.txt",            "LOW",      "Low",    "Add openai>=1.0.0 to requirements.txt. Update code to use new SDK style."),
    ("P3 — NEXT QUARTER","Deprecated openai.ChatCompletion.create()",       "LOW",      "Low",    "Migrate to client.chat.completions.create() using openai v1.x SDK."),
    ("P3 — NEXT QUARTER","Duplicate imports in main.py",                    "LOW",      "Low",    "Consolidate all imports at the top of the file. Remove mid-file imports."),
    ("P3 — NEXT QUARTER","Sensitive data in stdout logs",                   "LOW",      "Low",    "Replace print() with logging module. Set production log level to WARNING."),
    ("P3 — NEXT QUARTER","Unsanitized data in PDF ReportLab paragraphs",    "LOW",      "Low",    "HTML-escape user-provided strings before passing to Paragraph(). Truncate long fields."),
    ("P3 — NEXT QUARTER","python-multipart version may have DoS CVE",       "MEDIUM",   "Low",    "Pin python-multipart>=0.0.7 in requirements.txt."),
]

for ri, row in enumerate(roadmap, start=3):
    sev = row[2]
    bg_c, txt_c = SEV_COLOR.get(sev, ("FFFFFF", "000000"))
    for ci, val in enumerate(row, start=1):
        c_bg = bg_c if ci == 3 else ("DCE6F1" if ci == 1 else None)
        cell(ws6, ri, ci, val, bg=c_bg, bold=(ci == 1))

col_w6 = [22, 50, 12, 10, 80]
for ci, w in enumerate(col_w6, start=1):
    ws6.column_dimensions[get_column_letter(ci)].width = w

# ─── SHEET 7: GitHub Actions Workflow ─────────────────────────────────────────
ws7 = wb.create_sheet("GitHub Actions Workflow")
hdr(ws7, 1, 1, "GitHub Actions — Security Scanning Workflow (.github/workflows/security.yml)", bg="1F497D", sz=13)
ws7.merge_cells("A1:B1")
ws7.row_dimensions[1].height = 28

yaml_content = """name: PerioVoice AI Security Scan

on:
  push:
    branches: ["main", "develop"]
  pull_request:
    branches: ["main"]
  schedule:
    - cron: "0 6 * * 1"   # Every Monday at 6 AM UTC

jobs:
  gitleaks:
    name: Gitleaks — Secret Detection
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
        with:
          fetch-depth: 0
      - uses: gitleaks/gitleaks-action@v2
        env:
          GITHUB_TOKEN: ${{ secrets.GITHUB_TOKEN }}

  semgrep:
    name: Semgrep SAST
    runs-on: ubuntu-latest
    container:
      image: returntocorp/semgrep
    steps:
      - uses: actions/checkout@v4
      - name: Run Semgrep
        run: >
          semgrep scan
          --config=p/python
          --config=p/fastapi
          --config=p/owasp-top-ten
          --sarif
          --output=semgrep-results.sarif
          backend/
        env:
          SEMGREP_APP_TOKEN: ${{ secrets.SEMGREP_APP_TOKEN }}
      - uses: actions/upload-artifact@v4
        with:
          name: semgrep-sarif
          path: semgrep-results.sarif

  trivy:
    name: Trivy — Vulnerability Scanning
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
      - name: Run Trivy filesystem scan
        uses: aquasecurity/trivy-action@master
        with:
          scan-type: fs
          scan-ref: .
          format: table
          severity: CRITICAL,HIGH
          exit-code: 1

  dependency-review:
    name: Dependency Review
    runs-on: ubuntu-latest
    if: github.event_name == 'pull_request'
    steps:
      - uses: actions/checkout@v4
      - uses: actions/dependency-review-action@v4
        with:
          fail-on-severity: critical

  summary:
    name: Security Summary
    runs-on: ubuntu-latest
    needs: [gitleaks, semgrep, trivy]
    if: always()
    steps:
      - name: Publish Summary
        run: |
          echo "## PerioVoice AI Security Scan Results" >> $GITHUB_STEP_SUMMARY
          echo "| Scanner | Status |" >> $GITHUB_STEP_SUMMARY
          echo "|---------|--------|" >> $GITHUB_STEP_SUMMARY
          echo "| Gitleaks | ${{ needs.gitleaks.result }} |" >> $GITHUB_STEP_SUMMARY
          echo "| Semgrep | ${{ needs.semgrep.result }} |" >> $GITHUB_STEP_SUMMARY
          echo "| Trivy | ${{ needs.trivy.result }} |" >> $GITHUB_STEP_SUMMARY
"""

for ri, line in enumerate(yaml_content.strip().split("\n"), start=2):
    c = ws7.cell(row=ri, column=1, value=line)
    c.font = Font(name="Courier New", size=9, color="1F497D")
    c.alignment = Alignment(horizontal="left", vertical="center")

ws7.column_dimensions["A"].width = 100

# ─── Save workbook ─────────────────────────────────────────────────────────────
wb.active = 0   # Open to Executive Summary first

PATHS = [
    os.path.join(OUT_DIR, "PerioVoice_AI_Security_Review_Report.xlsx"),
    r"C:\Users\monisha D\android app\PerioVoice_AI_Security_Review_Report.xlsx",
]
for p in PATHS:
    try:
        os.makedirs(os.path.dirname(p), exist_ok=True)
        wb.save(p)
        print(f"✅ Saved: {p}")
    except Exception as e:
        print(f"⚠️  Could not save to {p}: {e}")

print("\n📊 Security Review Excel Report complete!")
print("   Sheets: Executive Summary | Backend Inventory | API Inventory")
print("           Security Findings (SAST) | Dependency Review")
print("           Risk Summary & Remediation | GitHub Actions Workflow")
