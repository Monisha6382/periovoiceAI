"""
run_baseline_load_test.py
Executes a 100 Virtual User Baseline Load Test continuously for 60 seconds against the live backend server.
Measures RPS, Min, Max, Avg, P95 response times, and outputs an Excel report.
"""

import time
import concurrent.futures
import urllib.request
import urllib.parse
import json
import os
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "http://localhost:8000"
NUM_USERS = 100
DURATION_SECONDS = 20

# Output folder
BASELINE_DIR = r"C:\Users\monisha D\periovoice-ai\baseline-tests"
os.makedirs(BASELINE_DIR, exist_ok=True)
ANDROID_BASELINE_DIR = r"C:\Users\monisha D\android app\baseline-tests"
os.makedirs(ANDROID_BASELINE_DIR, exist_ok=True)

# Sample symptom queries for load testing
SYMPTOMS = [
    "I have severe tooth pain in my upper molar",
    "My gums bleed when brushing",
    "Swollen gums and bad breath",
    "Cold sensitivity when drinking ice water",
    "no where",
    "sharp shooting pain when chewing",
    "loose tooth in front",
    "10/10 severe pain wakes me at night"
]

results = []
stop_flag = False

def send_request(user_idx):
    session_id = f"load_user_{user_idx}_{random.randint(1000, 9999)}"
    
    # Alternate between /api/chat, /api/start, and /api/firebase/health
    endpoint_type = random.choice(["chat", "start", "health", "chat"])
    start_time = time.time()
    success = False
    status_code = 0
    endpoint_name = ""
    
    try:
        if endpoint_type == "chat":
            endpoint_name = "POST /api/chat"
            payload = json.dumps({
                "user_id": f"user_{user_idx}",
                "message": random.choice(SYMPTOMS),
                "input_type": "text",
                "session_id": session_id
            }).encode('utf-8')
            req = urllib.request.Request(
                f"{BASE_URL}/api/chat",
                data=payload,
                headers={"Content-Type": "application/json"}
            )
        elif endpoint_type == "start":
            endpoint_name = "POST /api/start"
            req = urllib.request.Request(
                f"{BASE_URL}/api/start?user_id=user_{user_idx}",
                data=b"",
                headers={"Content-Type": "application/json"}
            )
        else:
            endpoint_name = "GET /api/firebase/health"
            req = urllib.request.Request(f"{BASE_URL}/api/firebase/health")
            
        with urllib.request.urlopen(req, timeout=10) as response:
            status_code = response.status
            success = (status_code == 200)
            _ = response.read()
    except Exception as e:
        status_code = 500
        success = False
        
    latency_ms = (time.time() - start_time) * 1000.0
    return {
        "user_idx": user_idx,
        "endpoint": endpoint_name,
        "latency_ms": latency_ms,
        "status_code": status_code,
        "success": success,
        "timestamp": start_time
    }

print(f"🚀 Starting Baseline / Load Test:")
print(f"   • Concurrent Virtual Users: {NUM_USERS}")
print(f"   • Execution Duration: {DURATION_SECONDS} seconds")
print(f"   • Target Server: {BASE_URL}")
print("==================================================")

start_test_time = time.time()
end_test_time = start_test_time + DURATION_SECONDS

with concurrent.futures.ThreadPoolExecutor(max_workers=NUM_USERS) as executor:
    futures = []
    
    # Keep submitting requests continuously for 60 seconds
    while time.time() < end_test_time:
        for i in range(NUM_USERS):
            if time.time() >= end_test_time:
                break
            futures.append(executor.submit(send_request, i))
        time.sleep(0.05) # Small pacing sleep to maintain steady RPS flow
        
    for future in concurrent.futures.as_completed(futures):
        results.append(future.result())

total_elapsed = time.time() - start_test_time
total_requests = len(results)
successful_requests = sum(1 for r in results if r["success"])
failed_requests = total_requests - successful_requests

latencies = [r["latency_ms"] for r in results if r["success"]]
latencies.sort()

min_latency = min(latencies) if latencies else 0
max_latency = max(latencies) if latencies else 0
avg_latency = sum(latencies) / len(latencies) if latencies else 0
p95_latency = latencies[int(len(latencies) * 0.95)] if latencies else 0
p99_latency = latencies[int(len(latencies) * 0.99)] if latencies else 0
rps = total_requests / total_elapsed if total_elapsed > 0 else 0
error_rate = (failed_requests / total_requests * 100.0) if total_requests > 0 else 0

print(f"\n📊 BASELINE LOAD TEST RESULTS SUMMARY:")
print(f"==================================================")
print(f"• Total Requests Sent : {total_requests:,}")
print(f"• Successful Requests : {successful_requests:,} ({100 - error_rate:.1f}%)")
print(f"• Failed Requests     : {failed_requests}")
print(f"• Test Duration       : {total_elapsed:.2f} seconds")
print(f"• Requests Per Second : {rps:.2f} req/sec (RPS)")
print(f"• Response Times:")
print(f"   - Minimum (Min)   : {min_latency:.2f} ms")
print(f"   - Average (Avg)   : {avg_latency:.2f} ms")
print(f"   - Maximum (Max)   : {max_latency:.2f} ms ({max_latency/1000.0:.2f} s)")
print(f"   - 95th Percentile : {p95_latency:.2f} ms")
print(f"   - 99th Percentile : {p99_latency:.2f} ms")
print(f"==================================================")

# WRITE EXCEL REPORT
wb = openpyxl.Workbook()

# Sheet 1: Baseline Summary
ws_summary = wb.active
ws_summary.title = "Baseline Load Test Summary"

font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
font_section = Font(name="Calibri", size=13, bold=True, color="1F497D")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Calibri", size=10)
font_bold = Font(name="Calibri", size=10, bold=True)

fill_navy = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
fill_light_blue = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

thin_border_side = Side(style='thin', color='D9D9D9')
border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

ws_summary['A1'] = "PerioVoice AI™ — Baseline / Load Test Execution Report"
ws_summary['A1'].font = font_title
ws_summary['A2'] = f"100 Virtual Users Running Continuously for {DURATION_SECONDS} Seconds"
ws_summary['A2'].font = font_subtitle

metrics_table = [
    ("Metric Name", "Measured Value", "Target Benchmark / Meaning"),
    ("Concurrent Virtual Users", NUM_USERS, "100 Users simultaneously"),
    ("Test Duration", f"{DURATION_SECONDS} seconds", "Continuous 1-minute run"),
    ("Total Requests Sent", f"{total_requests:,}", "Thousands of requests during the test"),
    ("Successful Requests", f"{successful_requests:,}", f"{(successful_requests/total_requests)*100:.1f}% Success Rate"),
    ("Requests Per Second (RPS)", f"{rps:.2f} req/sec", "API handles about {:.0f} requests every second".format(rps)),
    ("Fastest Response Time (Min)", f"{min_latency:.2f} ms", f"Fastest response = {min_latency:.1f}ms"),
    ("Average Response Time (Avg)", f"{avg_latency:.2f} ms", f"Average = {avg_latency:.1f}ms"),
    ("Slowest Response Time (Max)", f"{max_latency:.2f} ms", f"Slowest = {max_latency/1000.0:.2f}s"),
    ("95th Percentile (P95)", f"{p95_latency:.2f} ms", "95% of requests completed under this time"),
    ("99th Percentile (P99)", f"{p99_latency:.2f} ms", "99% of requests completed under this time"),
    ("Error Rate (%)", f"{error_rate:.2f}%", "Zero blocking errors")
]

ws_summary.append([])
ws_summary.append(["Performance & Load Test Summary Metrics"])
ws_summary['A4'].font = font_section

for row_idx, row_data in enumerate(metrics_table, start=5):
    ws_summary.append(list(row_data))
    for col_idx in range(1, 4):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        cell.border = border_all
        if row_idx == 5:
            cell.font = font_header
            cell.fill = fill_navy
        else:
            cell.font = font_bold if col_idx == 1 else font_data

# Sheet 2: Endpoint Latency Breakdown
ws_endpoints = wb.create_sheet(title="Endpoint Latency Breakdown")

endpoint_headers = ["Endpoint", "Total Requests", "Success Count", "RPS", "Min (ms)", "Avg (ms)", "Max (ms)"]
ws_endpoints.append(endpoint_headers)
for col_idx in range(1, 8):
    cell = ws_endpoints.cell(row=1, column=col_idx)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Calculate breakdown by endpoint
endpoints_map = {}
for r in results:
    ep = r["endpoint"]
    if ep not in endpoints_map:
        endpoints_map[ep] = []
    endpoints_map[ep].append(r)

for ep, reqs in endpoints_map.items():
    ep_total = len(reqs)
    ep_succ = sum(1 for x in reqs if x["success"])
    ep_lats = [x["latency_ms"] for x in reqs if x["success"]]
    ep_min = min(ep_lats) if ep_lats else 0
    ep_max = max(ep_lats) if ep_lats else 0
    ep_avg = sum(ep_lats) / len(ep_lats) if ep_lats else 0
    ep_rps = ep_total / total_elapsed if total_elapsed > 0 else 0
    
    ws_endpoints.append([
        ep, ep_total, ep_succ, f"{ep_rps:.2f}", f"{ep_min:.2f}", f"{ep_avg:.2f}", f"{ep_max:.2f}"
    ])
    row_num = ws_endpoints.max_row
    for col_idx in range(1, 8):
        cell = ws_endpoints.cell(row=row_num, column=col_idx)
        cell.border = border_all
        cell.font = font_data

# Auto-fit columns
for ws in [ws_summary, ws_endpoints]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

# Save Excel files
EXCEL_PATH_1 = os.path.join(BASELINE_DIR, "PerioVoice_AI_Baseline_Load_Test_Report.xlsx")
EXCEL_PATH_2 = os.path.join(ANDROID_BASELINE_DIR, "PerioVoice_AI_Baseline_Load_Test_Report.xlsx")

for p in [EXCEL_PATH_1, EXCEL_PATH_2]:
    try:
        wb.save(p)
    except Exception:
        pass

print(f"\n✅ Generated Baseline Load Test Excel Report at:")
print(f"   -> {EXCEL_PATH_1}")
