"""
PerioVoice AI™ - Automated Test Runner & Multi-Format Report Exporter
=====================================================================
Executes the 4 Core Automated Test Cases and generates comprehensive,
downloadable test report artifacts:
1. PerioVoice_AI_Test_Report.html (Interactive HTML Dashboard)
2. PerioVoice_AI_Automated_Test_Report.xlsx (Formatted Excel Test Matrix)
3. PerioVoice_Clinical_Assessment_Sample.pdf (Clinical PDF Export Sample)
4. test_summary.md (Markdown Executive Summary)
5. test_results.json (Structured JSON Metrics)

Outputs all files into the 'test-artifacts/' directory for GitHub Actions upload.
"""

import os
import sys
import time
import json
import traceback
from datetime import datetime

# Setup paths
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
ARTIFACTS_DIR = os.path.join(PROJECT_ROOT, "test-artifacts")
os.makedirs(ARTIFACTS_DIR, exist_ok=True)

if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)
if CURRENT_DIR not in sys.path:
    sys.path.insert(0, CURRENT_DIR)

# Import test cases
from backend.test_suite_4_cases import (
    test_case_1_triage_engine_and_guardrails,
    test_case_2_ai_clinical_urgency_assessment,
    test_case_3_image_analyzer_and_vision_ai,
    test_case_4_pdf_report_export_and_download
)
from backend.pdf_generator import pdf_generator
from backend.models import UrgencyLevel


def run_all_tests():
    """Execute the 4 core test cases and record results."""
    print("=" * 80)
    print("  🦷 PERIOVOICE AI™ - AUTOMATED CI TEST RUNNER (4 CORE TEST CASES)")
    print("=" * 80)
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}")
    print(f"Artifacts Destination: {ARTIFACTS_DIR}\n")

    test_definitions = [
        {
            "id": "TC-001",
            "name": "Triage State Engine & Adaptive Clinical Guardrails",
            "category": "Conversational AI & Guardrails",
            "description": "Verifies session initialization, off-topic query redirection, gibberish rejection without premature assessment, multi-turn clinical detail extraction, and emergency triage escalation.",
            "fn": test_case_1_triage_engine_and_guardrails
        },
        {
            "id": "TC-002",
            "name": "AI Clinical Urgency & Risk Assessment Calculation",
            "category": "Clinical Scoring & Reasoning",
            "description": "Verifies AI urgency tiers (LOW, MODERATE, HIGH, EMERGENCY), mathematical risk score bounds (0-10 scale), symptom mapping, and customized home-care tip generation.",
            "fn": test_case_2_ai_clinical_urgency_assessment
        },
        {
            "id": "TC-003",
            "name": "Oral Image Analyzer & Vision AI Clinical Classifier",
            "category": "Computer Vision & Diagnostics",
            "description": "Verifies dental image validation (dimensions/formats), rejection of non-dental photos (monochrome, software screens, documents), and clinical feature detection (erythema, swelling, calculus, pustules).",
            "fn": test_case_3_image_analyzer_and_vision_ai
        },
        {
            "id": "TC-004",
            "name": "Clinical PDF Assessment Report Export & Download Delivery",
            "category": "Reporting & Document Export",
            "description": "Verifies dynamic generation of patient periodontal assessment PDF reports across all 4 urgency tiers, structure integrity (%PDF- header), clinical disclaimers, and physical disk export.",
            "fn": test_case_4_pdf_report_export_and_download
        }
    ]

    results = []
    total_start = time.time()

    for idx, test in enumerate(test_definitions, start=1):
        print(f"[{idx}/4] Executing {test['id']}: {test['name']}...")
        t_start = time.time()
        status = "PASSED"
        error_msg = None

        try:
            test["fn"]()
            duration = round(time.time() - t_start, 3)
            print(f"      ✅ Status: PASSED (Duration: {duration}s)\n")
        except Exception as e:
            status = "FAILED"
            duration = round(time.time() - t_start, 3)
            error_msg = f"{str(e)}\n{traceback.format_exc()}"
            print(f"      ❌ Status: FAILED (Duration: {duration}s)\n      Error: {e}\n")

        results.append({
            "id": test["id"],
            "name": test["name"],
            "category": test["category"],
            "description": test["description"],
            "status": status,
            "duration_seconds": duration,
            "error": error_msg
        })

    total_duration = round(time.time() - total_start, 3)
    passed_count = sum(1 for r in results if r["status"] == "PASSED")
    failed_count = sum(1 for r in results if r["status"] == "FAILED")
    pass_rate = round((passed_count / len(results)) * 100, 1)

    summary = {
        "suite_name": "PerioVoice AI Core 4 Automated Test Suite",
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_tests": len(results),
        "passed": passed_count,
        "failed": failed_count,
        "pass_rate_percent": pass_rate,
        "total_duration_seconds": total_duration,
        "overall_status": "PASSED" if failed_count == 0 else "FAILED",
        "tests": results
    }

    # Generate Reports
    generate_json_report(summary)
    generate_markdown_summary(summary)
    generate_html_report(summary)
    generate_excel_report(summary)
    generate_sample_pdf_report()

    print("=" * 80)
    print(f"  🏁 SUMMARY: {passed_count}/{len(results)} PASSED ({pass_rate}%) in {total_duration}s")
    print(f"  📁 Artifacts generated in: {ARTIFACTS_DIR}")
    print("=" * 80)

    return failed_count == 0


def generate_json_report(summary: dict):
    """Write structured JSON results."""
    json_path = os.path.join(ARTIFACTS_DIR, "test_results.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    print(f"  -> Generated JSON Report: {json_path}")


def generate_markdown_summary(summary: dict):
    """Write Markdown summary for CI annotations."""
    md_path = os.path.join(ARTIFACTS_DIR, "test_summary.md")
    status_emoji = "🟢 PASS" if summary["overall_status"] == "PASSED" else "🔴 FAIL"

    lines = [
        f"# PerioVoice AI™ - Automated Test Suite Results\n",
        f"**Status:** {status_emoji} | **Pass Rate:** {summary['pass_rate_percent']}% | **Total Duration:** {summary['total_duration_seconds']}s | **Date:** {summary['timestamp']}\n",
        "## Test Execution Matrix\n",
        "| ID | Test Case | Category | Status | Duration |",
        "| :--- | :--- | :--- | :---: | :---: |"
    ]

    for t in summary["tests"]:
        badge = "✅ PASS" if t["status"] == "PASSED" else "❌ FAIL"
        lines.append(f"| **{t['id']}** | {t['name']} | {t['category']} | {badge} | {t['duration_seconds']}s |")

    lines.append("\n## Downloadable Artifacts")
    lines.append("- `PerioVoice_AI_Test_Report.html` - Interactive UI Dashboard")
    lines.append("- `PerioVoice_AI_Automated_Test_Report.xlsx` - Excel Test Case Matrix")
    lines.append("- `PerioVoice_Clinical_Assessment_Sample.pdf` - Sample Patient Assessment PDF")
    lines.append("- `test_results.json` - Machine Readable Execution Data")
    lines.append("\n*Generated automatically by PerioVoice AI CI Pipeline.*")

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  -> Generated Markdown Summary: {md_path}")


def generate_html_report(summary: dict):
    """Generate high-end, responsive HTML test report."""
    html_path = os.path.join(ARTIFACTS_DIR, "PerioVoice_AI_Test_Report.html")

    status_color = "#10b981" if summary["overall_status"] == "PASSED" else "#ef4444"
    status_text = "ALL TESTS PASSED" if summary["overall_status"] == "PASSED" else "TESTS FAILED"

    test_rows_html = ""
    for t in summary["tests"]:
        badge_bg = "#ecfdf5" if t["status"] == "PASSED" else "#fef2f2"
        badge_text = "#065f46" if t["status"] == "PASSED" else "#991b1b"
        icon = "✓" if t["status"] == "PASSED" else "✗"

        test_rows_html += f"""
        <tr class="test-row">
            <td class="font-mono text-sm font-semibold" style="color: #4338ca;">{t['id']}</td>
            <td>
                <div class="font-bold text-gray-900">{t['name']}</div>
                <div class="text-xs text-gray-500 mt-1">{t['description']}</div>
            </td>
            <td><span class="category-tag">{t['category']}</span></td>
            <td>
                <span class="status-badge" style="background: {badge_bg}; color: {badge_text};">
                    {icon} {t['status']}
                </span>
            </td>
            <td class="text-right font-mono text-sm text-gray-600">{t['duration_seconds']}s</td>
        </tr>
        """

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PerioVoice AI™ - Automated Test Report</title>
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
    <style>
        :root {{
            --primary: #4f46e5;
            --primary-dark: #3730a3;
            --bg: #f8fafc;
            --card-bg: #ffffff;
            --border: #e2e8f0;
            --text-main: #0f172a;
            --text-muted: #64748b;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Plus Jakarta Sans', sans-serif;
            background-color: var(--bg);
            color: var(--text-main);
            padding: 32px 16px;
            line-height: 1.5;
        }}
        .container {{
            max-width: 1080px;
            margin: 0 auto;
        }}
        .header {{
            background: linear-gradient(135deg, #1e1b4b 0%, #312e81 100%);
            border-radius: 20px;
            padding: 36px 40px;
            color: white;
            box-shadow: 0 10px 25px -5px rgba(49, 46, 129, 0.2);
            margin-bottom: 24px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 20px;
        }}
        .header h1 {{
            font-size: 26px;
            font-weight: 800;
            letter-spacing: -0.5px;
            display: flex;
            align-items: center;
            gap: 12px;
        }}
        .header-badge {{
            display: inline-flex;
            align-items: center;
            gap: 6px;
            background: {status_color};
            color: white;
            font-size: 13px;
            font-weight: 700;
            padding: 8px 16px;
            border-radius: 9999px;
            letter-spacing: 0.5px;
        }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .metric-card {{
            background: var(--card-bg);
            border: 1px solid var(--border);
            border-radius: 16px;
            padding: 20px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        }}
        .metric-label {{
            font-size: 13px;
            color: var(--text-muted);
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .metric-value {{
            font-size: 28px;
            font-weight: 800;
            margin-top: 6px;
            color: var(--text-main);
        }}
        .table-card {{
            background: var(--card-bg);
            border: 1px solid var(--border);
            border-radius: 16px;
            overflow: hidden;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05);
            margin-bottom: 24px;
        }}
        .table-header {{
            padding: 20px 24px;
            border-bottom: 1px solid var(--border);
            font-size: 17px;
            font-weight: 700;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
        }}
        th {{
            background: #f1f5f9;
            color: var(--text-muted);
            font-size: 12px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            padding: 14px 20px;
            border-bottom: 1px solid var(--border);
        }}
        td {{
            padding: 18px 20px;
            border-bottom: 1px solid var(--border);
            vertical-align: top;
        }}
        .status-badge {{
            display: inline-flex;
            align-items: center;
            gap: 4px;
            font-size: 12px;
            font-weight: 700;
            padding: 4px 10px;
            border-radius: 9999px;
        }}
        .category-tag {{
            background: #eef2ff;
            color: #4338ca;
            font-size: 11px;
            font-weight: 600;
            padding: 4px 8px;
            border-radius: 6px;
        }}
        .footer {{
            text-align: center;
            color: var(--text-muted);
            font-size: 13px;
            padding: 16px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div>
                <h1>🦷 PerioVoice AI™ Automated Test Report</h1>
                <p style="color: #c7d2fe; font-size: 14px; margin-top: 6px;">Continuous Integration & Clinical Diagnostic Test Suite</p>
            </div>
            <div class="header-badge">
                ● {status_text}
            </div>
        </div>

        <div class="metrics-grid">
            <div class="metric-card">
                <div class="metric-label">Total Test Cases</div>
                <div class="metric-value">{summary['total_tests']}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Passed Tests</div>
                <div class="metric-value" style="color: #10b981;">{summary['passed']}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Pass Rate</div>
                <div class="metric-value" style="color: #4f46e5;">{summary['pass_rate_percent']}%</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Total Execution Time</div>
                <div class="metric-value">{summary['total_duration_seconds']}s</div>
            </div>
        </div>

        <div class="table-card">
            <div class="table-header">
                <span>Core Automated Test Cases</span>
                <span style="font-size: 13px; font-weight: 500; color: var(--text-muted);">Timestamp: {summary['timestamp']} UTC</span>
            </div>
            <table>
                <thead>
                    <tr>
                        <th style="width: 90px;">Test ID</th>
                        <th>Test Case & Description</th>
                        <th style="width: 220px;">Category</th>
                        <th style="width: 120px;">Status</th>
                        <th style="width: 100px; text-align: right;">Duration</th>
                    </tr>
                </thead>
                <tbody>
                    {test_rows_html}
                </tbody>
            </table>
        </div>

        <div class="footer">
            PerioVoice AI™ Automated CI Suite • GitHub Actions Workflow Runner • Monisha6382/periovoiceAI
        </div>
    </div>
</body>
</html>
"""

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"  -> Generated HTML Report: {html_path}")


def generate_excel_report(summary: dict):
    """Generate professional Excel test matrix using openpyxl."""
    excel_path = os.path.join(ARTIFACTS_DIR, "PerioVoice_AI_Automated_Test_Report.xlsx")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CI Test Summary"

        # Styling definitions
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        font_pass = Font(name="Calibri", size=11, bold=True, color="065F46")
        font_fail = Font(name="Calibri", size=11, bold=True, color="991B1B")

        fill_title = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
        fill_header = PatternFill(start_color="3730A3", end_color="3730A3", fill_type="solid")
        fill_pass = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        # Title block
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "🦷 PerioVoice AI™ - Automated CI Test Execution Report"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Metadata block
        meta_info = [
            ("Execution Date:", summary["timestamp"] + " UTC"),
            ("Total Test Cases:", summary["total_tests"]),
            ("Passed Tests:", summary["passed"]),
            ("Pass Rate:", f"{summary['pass_rate_percent']}%"),
            ("Total Duration:", f"{summary['total_duration_seconds']} seconds"),
            ("Overall Status:", summary["overall_status"])
        ]

        for r_idx, (k, v) in enumerate(meta_info, start=3):
            ws.cell(row=r_idx, column=1, value=k).font = font_bold
            val_cell = ws.cell(row=r_idx, column=2, value=v)
            val_cell.font = font_regular
            if k == "Overall Status:":
                val_cell.font = font_pass if v == "PASSED" else font_fail

        # Headers
        headers = ["Test ID", "Test Case Name", "Category", "Expected Scope & Steps", "Status", "Duration (s)"]
        header_row = 10
        ws.row_dimensions[header_row].height = 25

        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col_idx, value=h)
            c.font = font_header
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Test Rows
        for idx, t in enumerate(summary["tests"], start=11):
            ws.row_dimensions[idx].height = 30
            is_alt = (idx % 2 == 0)

            c_id = ws.cell(row=idx, column=1, value=t["id"])
            c_name = ws.cell(row=idx, column=2, value=t["name"])
            c_cat = ws.cell(row=idx, column=3, value=t["category"])
            c_desc = ws.cell(row=idx, column=4, value=t["description"])
            c_status = ws.cell(row=idx, column=5, value=t["status"])
            c_dur = ws.cell(row=idx, column=6, value=t["duration_seconds"])

            c_id.font = font_bold
            c_name.font = font_bold
            c_cat.font = font_regular
            c_desc.font = font_regular
            c_dur.font = font_regular

            c_id.alignment = Alignment(horizontal="center", vertical="center")
            c_cat.alignment = Alignment(horizontal="center", vertical="center")
            c_status.alignment = Alignment(horizontal="center", vertical="center")
            c_dur.alignment = Alignment(horizontal="right", vertical="center")

            if t["status"] == "PASSED":
                c_status.fill = fill_pass
                c_status.font = font_pass
            else:
                c_status.fill = fill_fail
                c_status.font = font_fail

            for c in [c_id, c_name, c_cat, c_desc, c_status, c_dur]:
                c.border = thin_border
                if is_alt and c != c_status:
                    c.fill = fill_alt

        # Adjust column widths
        col_widths = {1: 12, 2: 40, 3: 30, 4: 55, 5: 14, 6: 16}
        for c_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        wb.save(excel_path)
        print(f"  -> Generated Excel Matrix: {excel_path}")
    except Exception as e:
        print(f"  ⚠️ Could not generate Excel report: {e}")


def generate_sample_pdf_report():
    """Generate a sample clinical PDF report demonstrating downloadable export."""
    pdf_path = os.path.join(ARTIFACTS_DIR, "PerioVoice_Clinical_Assessment_Sample.pdf")
    try:
        sample_payload = {
            "user_name": "Jane Smith (CI Sample)",
            "date": datetime.now().strftime("%Y-%m-%d"),
            "urgency_level": UrgencyLevel.HIGH,
            "risk_score": 7,
            "symptoms_found": ["bleeding_gums_brushing", "severe_swelling", "bad_breath_halitosis"],
            "recommendation": "Urgent dental evaluation recommended within 24-48 hours. Professional periodontal scaling and localized irrigation indicated.",
            "home_care_tips": [
                "Gently rinse with warm salt water (1/2 tsp salt in 8 oz water) 3 times daily.",
                "Use an ultra-soft toothbrush and avoid abrasive pressure on gingival margins.",
                "Maintain hydration and avoid tobacco or alcohol-based rinses."
            ],
            "detected_from_image": "Marked marginal erythema (redness) and localized gingival edema detected via Vision AI scanner.",
            "conversation_transcript": [
                {"isUser": True, "text": "My gums are sore, bleeding heavily when brushing and swollen"},
                {"isUser": False, "text": "How long have you experienced these symptoms?"},
                {"isUser": True, "text": "About 4 days now, pain is 7/10"}
            ]
        }
        pdf_generator.save_report(sample_payload, pdf_path)
        print(f"  -> Generated Sample PDF Report: {pdf_path}")
    except Exception as e:
        print(f"  ⚠️ Could not generate sample PDF report: {e}")


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
